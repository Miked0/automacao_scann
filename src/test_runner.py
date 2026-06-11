"""
M5 — TestRunner
Orquestra a validação linha a linha do roteiro XLSX.

Para cada linha executa 10 checks em sequência:
  1. Cupom localizado
  2. HTTP 200
  3. Cancelamento
  4. Total
  5. Desconto
  6. Meio de pagamento
  7. BIN
  8. Promoção (dispatcher PromoEngine)
  9. Desconto manual indevido
  10. Schema JSON mínimo
"""

from __future__ import annotations
import logging
import re
from typing import Any, Dict, List, Optional

from openpyxl.workbook import Workbook

from .models import AuditMovement, CheckResult, CouponBlock, TestResult
from .promo_engine import PromoEngine

log = logging.getLogger(__name__)

TOLERANCE = 0.05

# Mapeamento meio de pagamento → código Scanntech
_PAGTO_MAP = {
    "dinheiro": ["1", "01", "cash", "dinheiro"],
    "credito":  ["2", "02", "credit", "credito", "crédito", "cartao credito"],
    "debito":   ["3", "03", "debit", "debito",  "débito",  "cartao debito"],
    "pix":      ["4", "04", "pix"],
}

# Palavras-chave para detecção dinâmica do cabeçalho
_HEADER_KEYWORDS = {
    "teste", "tipo", "promo", "cupom", "total", "desconto",
    "pagamento", "bin", "observ", "resultado",
}


class TestRunner:
    """Itera sobre o roteiro e executa os 10 checks por linha."""

    def __init__(
        self,
        wb_roteiro: Workbook,
        movements:  Dict[str, AuditMovement],
        coupons:    List[CouponBlock],
        promo_engine: PromoEngine,
        extra_jsons: List[Dict] = None,
    ) -> None:
        self.wb          = wb_roteiro
        self.movements   = movements
        self.coupons     = coupons
        self.engine      = promo_engine
        self.extra_jsons = extra_jsons or []

    # ── Ponto de entrada ─────────────────────────────────────────────────────
    def run(self) -> List[TestResult]:
        results: List[TestResult] = []
        ws = self._get_main_sheet()
        header_row, col_map = self._detect_header(ws)
        if not col_map:
            log.error("Cabeçalho não detectado no roteiro. Abortando.")
            return results

        etapa = ws.title
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = self._extract_row(ws, row_idx, col_map)
            if not any(v for v in row_data.values() if v is not None and str(v).strip()):
                continue  # Linha vazia

            result = self._validate_row(etapa, row_idx, row_data)
            results.append(result)

        return results

    # ── Detecção de cabeçalho ────────────────────────────────────────────────
    def _get_main_sheet(self):
        for name in self.wb.sheetnames:
            if any(k in name.lower() for k in ("roteiro", "teste", "planilha", "sheet")):
                return self.wb[name]
        return self.wb.active

    def _detect_header(self, ws) -> tuple:
        for row in ws.iter_rows():
            texts = [str(c.value or "").lower().strip() for c in row]
            matches = sum(1 for t in texts if any(kw in t for kw in _HEADER_KEYWORDS))
            if matches >= 2:
                col_map = {}
                for cell in row:
                    key = str(cell.value or "").lower().strip()
                    col_map[key] = cell.column - 1  # 0-based
                return row[0].row, col_map
        return 1, {}

    def _extract_row(self, ws, row_idx: int, col_map: Dict) -> Dict:
        cells = [c.value for c in ws[row_idx]]
        return {key: (cells[idx] if idx < len(cells) else None)
                for key, idx in col_map.items()}

    # ── Validação de linha ───────────────────────────────────────────────────
    def _validate_row(self, etapa: str, row_idx: int, row: Dict) -> TestResult:
        result = TestResult(etapa=etapa, linha=row_idx)
        checks: List[CheckResult] = []

        cupom_key = self._normalize_cupom_key(row)
        result.cupom_key = cupom_key

        # 1. Cupom localizado
        movement, coupon_block = self._find_movement(cupom_key, row)
        if coupon_block:
            result.coupons_used.append(
                coupon_block.sat_number or coupon_block.nfce_number or coupon_block.coo_number or "?"
            )
        ck1 = CheckResult(
            check="cupom_localizado",
            ok=movement is not None,
            detalhe=f"Cupom '{cupom_key}' {'encontrado' if movement else 'NÃO encontrado'} no Audit.",
        )
        checks.append(ck1)

        if movement is None:
            result.checks      = checks
            result.overall_ok  = False
            result.error_reason = ck1.detalhe
            self._set_status_columns(result, "ERRO")
            return result

        # 2-10. Restantes checks
        checks.append(self._check_http200(movement))
        checks.append(self._check_cancelamento(row, movement))
        checks.append(self._check_total(row, movement))
        checks.append(self._check_desconto(row, movement))
        checks.append(self._check_pagamento(row, movement))
        checks.append(self._check_bin(row, movement))
        checks.append(self._check_promo(row, movement))
        checks.append(self._check_desconto_indevido(row, movement))
        checks.append(self._check_schema(movement))

        result.checks    = checks
        failed           = [c for c in checks if not c.ok]
        result.overall_ok = not failed
        if failed:
            result.error_reason = failed[0].detalhe[:100]

        self._set_status_columns(result, "Ok" if result.overall_ok else "Erro")
        return result

    # ── Checks individuais ───────────────────────────────────────────────────
    @staticmethod
    def _check_http200(mv: AuditMovement) -> CheckResult:
        ok = mv.status_code == 200
        return CheckResult(
            check="http200",
            ok=ok,
            detalhe=f"HTTP status={mv.status_code}",
        )

    @staticmethod
    def _check_cancelamento(row: Dict, mv: AuditMovement) -> CheckResult:
        obs = str(row.get("observ") or row.get("observacao") or "").lower()
        esperado_cancel = "cancel" in obs
        ok = mv.cancelacion == esperado_cancel
        return CheckResult(
            check="cancelamento",
            ok=ok,
            detalhe=f"cancelacion={mv.cancelacion} esperado={esperado_cancel}",
        )

    @staticmethod
    def _check_total(row: Dict, mv: AuditMovement) -> CheckResult:
        total_rot = _safe_float(row.get("total"))
        if total_rot is None:
            return CheckResult(check="total", ok=True, detalhe="Total não informado no roteiro — skip.")
        diff = abs((mv.total or 0.0) - total_rot)
        ok   = diff <= TOLERANCE
        return CheckResult(
            check="total",
            ok=ok,
            detalhe=f"total audit={mv.total:.2f} roteiro={total_rot:.2f} diff={diff:.4f}",
        )

    @staticmethod
    def _check_desconto(row: Dict, mv: AuditMovement) -> CheckResult:
        desc_rot = _safe_float(row.get("desconto"))
        if desc_rot is None:
            return CheckResult(check="desconto", ok=True, detalhe="Desconto não informado no roteiro — skip.")
        diff = abs((mv.descuento_total or 0.0) - desc_rot)
        ok   = diff <= TOLERANCE
        return CheckResult(
            check="desconto",
            ok=ok,
            detalhe=f"descuento audit={mv.descuento_total} roteiro={desc_rot} diff={diff:.4f}",
        )

    @staticmethod
    def _check_pagamento(row: Dict, mv: AuditMovement) -> CheckResult:
        tipo_rot = str(row.get("pagamento") or row.get("meio_pag") or "").lower().strip()
        if not tipo_rot:
            return CheckResult(check="pagamento", ok=True, detalhe="Pagamento não informado — skip.")

        codigos_aceitos = []
        for label, codigos in _PAGTO_MAP.items():
            if label in tipo_rot or tipo_rot in codigos:
                codigos_aceitos = codigos
                break

        for pago in mv.pagos:
            codigo = str(pago.get("codigoTipoPago") or "").lower()
            if codigo in codigos_aceitos or tipo_rot in codigos_aceitos:
                return CheckResult(check="pagamento", ok=True, detalhe=f"Pagamento '{tipo_rot}' OK.")

        return CheckResult(
            check="pagamento",
            ok=False,
            detalhe=f"Pagamento '{tipo_rot}' não encontrado nos pagos do Audit.",
        )

    @staticmethod
    def _check_bin(row: Dict, mv: AuditMovement) -> CheckResult:
        obs     = str(row.get("observ") or row.get("bin") or "").upper()
        bin_req = re.search(r"BIN[:\s]+([\dX]+)", obs)
        if not bin_req:
            return CheckResult(check="bin", ok=True, detalhe="BIN não exigido.")
        bin_esperado = bin_req.group(1)
        bin_audit    = str(mv.bin_value or "").upper()
        ok = bin_esperado in bin_audit
        return CheckResult(
            check="bin",
            ok=ok,
            detalhe=f"BIN esperado={bin_esperado} audit={bin_audit}",
        )

    def _check_promo(self, row: Dict, mv: AuditMovement) -> CheckResult:
        tipo_promo = str(row.get("tipo promo") or row.get("tipo_promo") or "").strip()
        if not tipo_promo:
            return CheckResult(check="promo", ok=True, detalhe="Tipo promoção não informado — skip.")
        return self.engine.validate(tipo_promo, row, mv)

    @staticmethod
    def _check_desconto_indevido(row: Dict, mv: AuditMovement) -> CheckResult:
        obs = str(row.get("observ") or "").lower()
        if "sem desconto manual" in obs:
            # Scanntech ativa: descuento deve ser somente o da promo
            desc_rot   = _safe_float(row.get("desconto")) or 0.0
            desc_audit = mv.descuento_total or 0.0
            excesso    = desc_audit - desc_rot
            ok = excesso <= TOLERANCE
            return CheckResult(
                check="desconto_indevido",
                ok=ok,
                detalhe=f"Desconto manual: excesso={excesso:.4f}",
            )
        return CheckResult(check="desconto_indevido", ok=True, detalhe="Sem restrição de desconto manual.")

    @staticmethod
    def _check_schema(mv: AuditMovement) -> CheckResult:
        obrigatorios = {"total", "numero", "detalles", "pagos"}
        presentes    = set(mv.raw_json.keys())
        faltando     = obrigatorios - presentes
        ok = not faltando
        return CheckResult(
            check="schema",
            ok=ok,
            detalhe=(
                "Schema OK."
                if ok
                else f"Campos ausentes: {', '.join(sorted(faltando))}"
            ),
        )

    # ── Busca de movimento ───────────────────────────────────────────────────
    def _find_movement(
        self, cupom_key: str, row: Dict
    ) -> tuple:
        """Busca por SAT/ECF/NFCE ou fallback por EANs."""
        if cupom_key and cupom_key in self.movements:
            return self.movements[cupom_key], None

        # Fallback: busca pelo conjunto de EANs do roteiro
        eans_rot = self._extract_eans_from_row(row)
        if eans_rot:
            for key, mv in self.movements.items():
                mv_eans = {str(d.get("codigoBarras") or "") for d in mv.detalles}
                if eans_rot.issubset(mv_eans):
                    log.debug("Match por EAN para cupom %s → %s", cupom_key, key)
                    return mv, None

        return None, None

    @staticmethod
    def _normalize_cupom_key(row: Dict) -> str:
        for campo in ("cupom", "nro_cupom", "numero_cupom", "sat", "nfce", "ecf"):
            val = row.get(campo)
            if val and str(val).strip():
                return str(val).strip().lower()
        return ""

    @staticmethod
    def _extract_eans_from_row(row: Dict) -> set:
        raw = str(row.get("ean") or row.get("eans") or "").strip()
        if not raw:
            return set()
        return {e.strip() for e in re.split(r"[,;\s]+", raw) if re.match(r"\d{7,14}", e.strip())}

    # ── Status columns ───────────────────────────────────────────────────────
    @staticmethod
    def _set_status_columns(result: TestResult, status: str) -> None:
        result.col_sat_status  = status
        result.col_ecf_status  = status
        result.col_nfce_status = status
        if status == "Erro":
            result.col_justificativa = result.error_reason[:100]


def _safe_float(val: Any) -> Optional[float]:
    try:
        return float(val)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
