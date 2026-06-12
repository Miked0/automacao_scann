"""
Gravador de resultados no TEMPLATE de roteiro de testes.

Responsabilidade: preencher as colunas de resultado (Ok/Erro) e salvar o xlsx.

Estrutura real do TEMPLATE (linha 7):
  C1  Teste        C6  SAT          C12-14 Json/Minoristas/Cupom (SAT)
  C2  Tipo Promo   C7  ECF          C15-17 Json/Minoristas/Cupom (ECF)
  C3  Itens        C8  NFCE         C18-20 Json/Minoristas/Cupom (NFCE)
  C4  Pagamento    C9  Sub-Total    C21    Observações (output)
  C5  Observações  C10 Desconto
                   C11 Total
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

# Cores de preenchimento para feedback visual no Excel
PREENCHIMENTO_APROVADO  = PatternFill("solid", fgColor="C6EFCE")  # verde
PREENCHIMENTO_REPROVADO = PatternFill("solid", fgColor="FFC7CE")  # vermelho
PREENCHIMENTO_NEUTRO    = PatternFill("solid", fgColor="FFEB9C")  # amarelo
PREENCHIMENTO_LIMPO     = PatternFill()                           # sem cor

PALAVRAS_CABECALHO_OBRIGATORIAS = {"teste", "promo", "pagamento"}
PALAVRAS_CABECALHO_OPCIONAIS    = {
    "desconto", "total", "cupom", "itens", "sub-total", "subtotal", "observ"
}

COLUNAS_GRUPOS_FIXOS: Dict[str, tuple] = {
    "SAT":  (12, 13, 14),
    "ECF":  (15, 16, 17),
    "NFCE": (18, 19, 20),
}
COLUNA_OBSERVACOES_SAIDA = 21

SUBCOLUNAS_RESULTADO = ("json", "minoristas", "cupom", "cup")


def detectar_linha_cabecalho(planilha) -> int:
    """
    Detecta a linha de cabeçalho exigindo ao menos 2 palavras-chave obrigatórias
    na mesma linha — evita falsos positivos em linhas de título ou dados (BUG-02).
    """
    melhor_linha, melhor_pontuacao = 1, 0

    for linha in planilha.iter_rows():
        valores_linha = {str(celula.value).lower() for celula in linha if celula.value}

        acertos_obrigatorios = sum(
            1 for palavra in PALAVRAS_CABECALHO_OBRIGATORIAS
            if any(palavra in valor for valor in valores_linha)
        )
        acertos_opcionais = sum(
            1 for palavra in PALAVRAS_CABECALHO_OPCIONAIS
            if any(palavra in valor for valor in valores_linha)
        )
        pontuacao = acertos_obrigatorios * 10 + acertos_opcionais

        if acertos_obrigatorios >= 2 and pontuacao > melhor_pontuacao:
            melhor_pontuacao = pontuacao
            melhor_linha = linha[0].row

    logger.info("Cabeçalho detectado na linha %d (pontuação=%d)", melhor_linha, melhor_pontuacao)
    return melhor_linha


def _localizar_coluna_por_palavras(
    planilha, linha_cabecalho: int, palavras_chave: List[str]
) -> Optional[int]:
    """Retorna o índice da coluna cujo cabeçalho contenha alguma das palavras-chave."""
    for celula in planilha[linha_cabecalho]:
        if celula.value:
            valor = str(celula.value).lower()
            if any(palavra.lower() in valor for palavra in palavras_chave):
                return celula.column
    return None


class GravadorResultados:
    """
    Preenche as colunas de resultado (Ok/Erro) nos 3 grupos SAT/ECF/NFCE
    e salva o arquivo xlsx ao final.
    """

    def __init__(self, pasta_trabalho: openpyxl.Workbook, caminho_saida: str):
        self.pasta_trabalho = pasta_trabalho
        self.caminho_saida  = Path(caminho_saida)
        self.planilha       = pasta_trabalho.active

        self.linha_cabecalho = detectar_linha_cabecalho(self.planilha)
        self.grupos          = self._detectar_grupos_resultado()
        self.coluna_observacoes = self._detectar_coluna_observacoes()

        logger.info(
            "GravadorResultados pronto | cabeçalho=linha %d | grupos=%s | obs=col %d",
            self.linha_cabecalho, self.grupos, self.coluna_observacoes,
        )

    def _detectar_grupos_resultado(self) -> Dict[str, List[int]]:
        """
        Tenta detectar os grupos SAT/ECF/NFCE dinamicamente pela linha acima do cabeçalho.
        Usa colunas fixas como fallback caso a detecção dinâmica falhe.
        """
        linha_grupo = self.linha_cabecalho - 1
        grupos: Dict[str, List[int]] = {"SAT": [], "ECF": [], "NFCE": []}

        if linha_grupo >= 1:
            grupo_atual = None
            for celula in self.planilha[linha_grupo]:
                if celula.value:
                    valor = str(celula.value).upper()
                    for nome_grupo in ("SAT", "ECF", "NFCE"):
                        if nome_grupo in valor:
                            grupo_atual = nome_grupo
                            break
                if grupo_atual:
                    subcoluna = self.planilha.cell(
                        row=self.linha_cabecalho, column=celula.column
                    ).value
                    if subcoluna and str(subcoluna).lower() in SUBCOLUNAS_RESULTADO:
                        grupos[grupo_atual].append(celula.column)

        todos_detectados = all(len(cols) == 3 for cols in grupos.values())
        if not todos_detectados:
            # Detecção dinâmica falhou — estrutura do TEMPLATE diferente do esperado
            logger.warning("Grupos não detectados dinamicamente. Usando colunas fixas.")
            return {nome: list(cols) for nome, cols in COLUNAS_GRUPOS_FIXOS.items()}

        return grupos

    def _detectar_coluna_observacoes(self) -> int:
        ultima_coluna_resultado = max(
            max(colunas) for colunas in self.grupos.values()
        )
        for indice_col in range(ultima_coluna_resultado + 1, ultima_coluna_resultado + 4):
            celula = self.planilha.cell(row=self.linha_cabecalho, column=indice_col)
            if celula.value and "obs" in str(celula.value).lower():
                return indice_col
        return COLUNA_OBSERVACOES_SAIDA

    def gravar_resultado(self, resultado, linha_dados: int) -> None:
        """Preenche as células de resultado com Ok/Erro e aplica cor de fundo."""
        texto_status  = "Ok" if resultado.passed else "Erro"
        preenchimento = PREENCHIMENTO_APROVADO if resultado.passed else PREENCHIMENTO_REPROVADO

        for colunas in self.grupos.values():
            for coluna in colunas:
                celula = self.planilha.cell(row=linha_dados, column=coluna)
                celula.value = texto_status
                celula.fill  = preenchimento

        celula_obs = self.planilha.cell(row=linha_dados, column=self.coluna_observacoes)
        celula_obs.value = resultado.motivo_erro[:100] if not resultado.passed else ""

    def limpar_linha(self, linha_dados: int) -> None:
        """Remove valores e formatação das células de resultado de uma linha."""
        for colunas in self.grupos.values():
            for coluna in colunas:
                celula = self.planilha.cell(row=linha_dados, column=coluna)
                celula.value = None
                celula.fill  = PREENCHIMENTO_LIMPO

        celula_obs = self.planilha.cell(row=linha_dados, column=self.coluna_observacoes)
        celula_obs.value = None
        celula_obs.fill  = PREENCHIMENTO_LIMPO

    def salvar(self) -> None:
        self.pasta_trabalho.save(self.caminho_saida)
        logger.info("Resultado salvo em: %s", self.caminho_saida)


# Alias de compatibilidade — mantido enquanto __init__.py referencia ResultWriter
ResultWriter = GravadorResultados
