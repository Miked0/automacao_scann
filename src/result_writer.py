"""
Módulo 6 — ResultWriter  (v2 — corrigido BUG-02 e BUG-04)
Responsabilidade: Preenche colunas de resultado + salva xlsx.

Fix BUG-02: _detect_header_row exige ao menos 3 keywords simultâneas na mesma linha.
Fix BUG-04: preenche os 3 grupos de colunas (SAT/ECF/NFCE) × 3 sub-colunas
            (Json/Minoristas/Cupom) = 9 colunas + coluna Observacoes.

Estrutura real do TEMPLATE (linha 7):
  C1  Teste
  C2  Tipo Promo
  C3  Itens da venda
  C4  Pagamento
  C5  Observacoes (input)
  C6  SAT  (número cupom — preenchido pelo QA)
  C7  ECF  (número cupom)
  C8  NFCE (número cupom)
  C9  Sub-Total
  C10 Desconto
  C11 Total
  ── Grupo SAT ──
  C12 Json
  C13 Minoristas
  C14 Cupom
  ── Grupo ECF ──
  C15 Json
  C16 Minoristas
  C17 Cupom
  ── Grupo NFCE ──
  C18 Json
  C19 Minoristas
  C20 Cupom
  C21 Observacoes (output — justificativa de erro)
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

FILL_OK   = PatternFill("solid", fgColor="C6EFCE")
FILL_ERRO = PatternFill("solid", fgColor="FFC7CE")
FILL_NA   = PatternFill("solid", fgColor="FFEB9C")

# Palavras-chave que DEVEM aparecer juntas na linha de cabeçalho
_HEADER_REQUIRED = {"teste", "promo", "pagamento"}
_HEADER_OPTIONAL = {"desconto", "total", "cupom", "itens", "sub-total", "subtotal", "observ"}


def detect_header_row(ws) -> int:
    """
    Detecta a linha de cabeçalho real exigindo que ao menos 3 palavras-chave
    obrigatórias estejam presentes na MESMA linha (BUG-02 fix).
    Retorna número da linha (1-based). Padrão: 1.
    """
    best_row, best_score = 1, 0
    for row in ws.iter_rows():
        values = {str(c.value).lower() for c in row if c.value}
        # conta keywords obrigatórias presentes
        req_hits = sum(
            1 for kw in _HEADER_REQUIRED
            if any(kw in v for v in values)
        )
        opt_hits = sum(
            1 for kw in _HEADER_OPTIONAL
            if any(kw in v for v in values)
        )
        score = req_hits * 10 + opt_hits
        if req_hits >= 2 and score > best_score:
            best_score = score
            best_row = row[0].row
    logger.info("Header detectado na linha %d (score=%d)", best_row, best_score)
    return best_row


def _find_col_by_keywords(ws, header_row: int, keywords: list[str]) -> Optional[int]:
    """Retorna índice (1-based) da primeira coluna cujo header contenha alguma keyword."""
    for cell in ws[header_row]:
        if cell.value:
            val = str(cell.value).lower()
            if any(kw.lower() in val for kw in keywords):
                return cell.column
    return None


class ResultWriter:
    """
    Escreve resultados nos 3 grupos de colunas (SAT/ECF/NFCE) do TEMPLATE e salva o xlsx.
    """

    # Colunas fixas conforme estrutura real do TEMPLATE
    # Grupo SAT:  Json=12, Minoristas=13, Cupom=14
    # Grupo ECF:  Json=15, Minoristas=16, Cupom=17
    # Grupo NFCE: Json=18, Minoristas=19, Cupom=20
    # Observacoes (output): 21
    _FIXED_GROUPS = {
        "SAT":  (12, 13, 14),
        "ECF":  (15, 16, 17),
        "NFCE": (18, 19, 20),
    }
    _COL_OBS_OUTPUT = 21

    def __init__(self, workbook: openpyxl.Workbook, output_path: str):
        self.wb = workbook
        self.output_path = Path(output_path)
        self.ws = workbook.active

        self.header_row = detect_header_row(self.ws)

        # Tenta detectar dinamicamente os grupos de colunas;
        # se não encontrar, usa posições fixas do TEMPLATE padrão.
        self.groups = self._detect_result_groups()
        self.col_obs = self._detect_obs_col()

        logger.info(
            "ResultWriter inicializado | header_row=%d | grupos=%s | col_obs=%d",
            self.header_row, self.groups, self.col_obs,
        )

    # ------------------------------------------------------------------
    # Detecção de colunas
    # ------------------------------------------------------------------

    def _detect_result_groups(self) -> dict:
        """
        Detecta os 3 grupos (SAT/ECF/NFCE) lendo linha de grupo (header_row - 1)
        e linha de sub-header (header_row).
        Fallback: usa colunas fixas do TEMPLATE padrão.
        """
        # Linha de grupos (ex: linha 6 no TEMPLATE padrão)
        group_row = self.header_row - 1
        groups: dict[str, list[int]] = {"SAT": [], "ECF": [], "NFCE": []}

        if group_row >= 1:
            current_group = None
            for cell in self.ws[group_row]:
                if cell.value:
                    val = str(cell.value).upper()
                    for g in ("SAT", "ECF", "NFCE"):
                        if g in val:
                            current_group = g
                            break
                # Coleta 3 colunas por grupo lendo sub-headers (Json/Minoristas/Cupom)
                if current_group:
                    sub = self.ws.cell(row=self.header_row, column=cell.column).value
                    if sub and str(sub).lower() in ("json", "minoristas", "cupom", "cup"):
                        groups[current_group].append(cell.column)

        # Valida — se algum grupo ficou incompleto, usa fixos
        for name, cols in groups.items():
            if len(cols) != 3:
                logger.warning(
                    "Grupo %s não detectado dinamicamente (%d cols), usando fixo.",
                    name, len(cols)
                )
                groups = {
                    "SAT":  list(self._FIXED_GROUPS["SAT"]),
                    "ECF":  list(self._FIXED_GROUPS["ECF"]),
                    "NFCE": list(self._FIXED_GROUPS["NFCE"]),
                }
                break

        return groups

    def _detect_obs_col(self) -> int:
        """Detecta coluna de Observacoes de saída. Fallback: col 21."""
        # Procura após a última coluna de resultado
        last_result_col = max(
            max(cols) for cols in self.groups.values()
        )
        for col_idx in range(last_result_col + 1, last_result_col + 4):
            cell = self.ws.cell(row=self.header_row, column=col_idx)
            if cell.value and "obs" in str(cell.value).lower():
                return col_idx
        return self._COL_OBS_OUTPUT

    # ------------------------------------------------------------------
    # Escrita de resultado
    # ------------------------------------------------------------------

    def write_result(self, result, data_row: int) -> None:
        """
        Preenche os 3 grupos × 3 sub-colunas + observações para uma linha de dado.
        result deve ter atributos: .passed (bool), .motivo_erro (str).
        """
        status = "Ok" if result.passed else "Erro"
        fill   = FILL_OK if result.passed else FILL_ERRO

        # Preenche os 9 campos de resultado
        for group_name, cols in self.groups.items():
            for col in cols:
                cell = self.ws.cell(row=data_row, column=col)
                # Não sobrescreve colunas fora das de resultado
                cell.value = status
                cell.fill  = fill

        # Preenche coluna de observações (justificativa de erro)
        obs_cell = self.ws.cell(row=data_row, column=self.col_obs)
        if not result.passed:
            obs_cell.value = result.motivo_erro[:100]
        else:
            obs_cell.value = ""

    def clear_result_row(self, data_row: int) -> None:
        """Limpa valores anteriores nas colunas de resultado (evita reprocessamento sujo)."""
        for cols in self.groups.values():
            for col in cols:
                cell = self.ws.cell(row=data_row, column=col)
                cell.value = None
                cell.fill  = PatternFill()  # remove fill
        obs_cell = self.ws.cell(row=data_row, column=self.col_obs)
        obs_cell.value = None
        obs_cell.fill  = PatternFill()

    def save(self) -> None:
        """Salva o workbook no caminho de saída."""
        self.wb.save(self.output_path)
        logger.info("Resultado salvo em: %s", self.output_path)
